#-------------------------------------------------Program that obtains the Linear Programming result and exports it to Excel-------------
#----------------------------------------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------------------------------------

# Import required libraries
from pulp import *
import Pasar_excel as Ex  # Import Excel export module
import Lectura_de_datos as lect  # Import data reading module
import re  # Regular expression module
from openpyxl import Workbook  # Excel workbook module
from openpyxl.styles import Font, PatternFill, Border, Side  # Style modules
from openpyxl.utils import get_column_letter  # Utility function to convert column index to letter

#-----------------collect the data----------------
# Read data from CSV file using custom module
grupos_alumnos2, grupos_alumnos1 = lect.leer("Datos EHU.csv")

# --------------------------------Define the problem-----------------------------
problem = LpProblem("Preferred_Class_Schedules", LpMaximize)  # Define LP problem with maximization objective

#------------------------------Define the lists and dictionaries to be used later-------------------------
asignaturas = []  # List to store subjects
indice_asignatura = {}  # Dictionary to map subject indices to names
profesores = {}  # Dictionary to map professors to subjects they teach
asig_4 = {}  # Dictionary to store 4th-year subjects and their elective restrictions
curso = {}  # Dictionary to store subjects by course
asignaturas_ga = {}  # Dictionary to store subjects that are GAs
asignaturas_ga_no = []  # List to store non-GA subjects
curso_comun = {}  # Dictionary to store common courses
curso_opt = {}  # Dictionary to store elective courses
curso_total = {}  # Dictionary to store total courses
maña_tard = {}  # Dictionary to store morning and afternoon coefficients
dias_semana = {}  # Dictionary to store how many days per week each subject is taught
horas_dia = {}  # Dictionary to store how many hours per day each subject is taught
otro_curso = {}  # Dictionary for courses other than common and elective
preferencias_hora = {}  # Dictionary to store professors' hour preferences
preferencias_dia = {}  # Dictionary to store professors' day preferences
aula = [1, 2, 3, 4, 5, 6, 7]  # List of classrooms
horarios = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]  # List of periods
dias = [1, 2, 3, 4, 5]  # List of days
indice = 0  # Index variable

#----------------------------Organize the data into different lists and dictionaries---------------------------------
# First, make a list with all the GA subjects that do not go with those that do
for datos1 in grupos_alumnos2:
    for datos in datos1:
        if datos[6][2][2][-3:-1] == "GA":
            if int(datos[6][2][-1]) > 1:
                if datos[6][11] != "No":
                    asignaturas_ga_no.append(datos[6][2])
# First, create a dictionary for each subject, where the key is a number that encodes each subject
for datos1 in grupos_alumnos2:
    for datos in datos1:
        if datos[6][2] not in asignaturas_ga_no:
            indice_asignatura[round((datos[1] % 1) * 1000)] = datos[6][2]

# Now, store all the subjects in a list 
for datos1 in grupos_alumnos2:
    for datos in datos1:
        if datos[6][2] not in asignaturas:
            asignaturas.append(datos[6][2])

        # Additionally, in a dictionary, save the subjects of each professor in a list (The professor's name is the key)
        if datos[6][3] in profesores:
            profesores[datos[6][3]].append(datos[6][2])
        else:
            profesores[datos[6][3]] = [datos[6][2]]

        # For each subject, depending on whether it is in the morning or in the afternoon, define a list with the coefficients for each hour
        if datos[2] == "No":  # If it is in the morning
            # First, check if they are GA, as they should be first or last hour due to split sessions
            if datos[6][2][-3:-1] == "GA":
                if datos[6][2] not in maña_tard:
                    maña_tard[datos[6][2]] = {1: 11, 2: 5, 3: 5, 4: 10, 5: 1, 6: 0, 7: -8, 8: -8, 9: -10, 10: -11}
            else:
                if datos[6][2] not in maña_tard:
                    maña_tard[datos[6][2]] = {1: 11, 2: 11, 3: 10, 4: 9, 5: 1, 6: 0, 7: -8, 8: -8, 9: -10, 10: -11}

        else:  # If it is in the afternoon
            # First, check if they are GA, as they should be first or last hour due to split sessions
            if datos[6][2][-3:-1] == "GA":
                if datos[6][2] not in maña_tard:
                    maña_tard[datos[6][2]] = {1: -10, 2: -10, 3: -10, 4: -10, 5: -1, 6: 0, 7: 11, 8: 5, 9: 5, 10: 8}
            else:
                if datos[6][2] not in maña_tard:
                    maña_tard[datos[6][2]] = {1: -10, 2: -10, 3: -10, 4: -10, 5: -1, 6: 0, 7: 11, 8: 10, 9: 9, 10: 8}

        # Save how many days a week the subject has classes in a dictionary
        if datos[6][2] not in dias_semana:
            dias_semana[datos[6][2]] = int(datos[6][8])

        # Save how many hours per day the subject has classes in a dictionary
        if datos[6][2] not in horas_dia:
            horas_dia[datos[6][2]] = int(datos[6][7])

        # Save the preferences that professors have for hours
        if datos[6][3] not in preferencias_hora:
            preferencias_hora[datos[6][3]] = {}
            horas_pre = datos[6][9].split(" ")
            if len(horas_pre) > 0:
                for i in range(10):
                    for j in horas_pre:
                        if i + 1 == int(j):
                            preferencias_hora[datos[6][3]][i + 1] = 1.5
                        else:
                            preferencias_hora[datos[6][3]][i + 1] = 1

        # Save the preferences that professors have for days
        if datos[6][3] not in preferencias_dia:
            preferencias_dia[datos[6][3]] = {}
            dias_pre = datos[6][10].split(" ")
            if len(dias_pre) > 0:
                for i in range(5):
                    for j in dias_pre:
                        if i + 1 == int(j):
                            preferencias_dia[datos[6][3]][i + 1] = 1.5
                        else:
                            preferencias_dia[datos[6][3]][i + 1] = 1

        # Also, save the subjects that cannot coincide with subjects from another course
        if datos[6][2] in asignaturas_ga_no:
            nombre = datos[6][11]
            asignaturas_ga[nombre] = [nombre, datos[6][2]]
            continue

        # In a dictionary, save the subjects of each course
        if datos[6][1] != "Optativa":
            if int(datos[1]) in curso:
                curso[int(datos[1])].append(datos[6][2])
                curso_comun[int(datos[1])].append(datos[6][2])
            else:
                curso[int(datos[1])] = [datos[6][2]]
                curso_comun[int(datos[1])] = [datos[6][2]]
            if len(datos[5]) > 0:
                for asig in datos[5]:
                    if asig in indice_asignatura:
                        if indice_asignatura[asig] not in curso_comun[int(datos[1])]:
                            curso_comun[int(datos[1])].append(indice_asignatura[asig])
        else:  # If it's elective, store it by mention
            if datos[6][0] + "_" + datos[6][5] in curso_opt:
                curso_opt[datos[6][0]+"_"+datos[6][5]].append(datos[6][2])
            else:
                curso_opt[datos[6][0]+"_"+datos[6][5]] = [datos[6][2]]

        #To ensure that electives do not coincide with those of 4, we initialize different nested dictionaries.
        if datos[6][1] == "4":
            if datos[6][0] not in asig_4:
                asig_4[datos[6][0]] = {}
                if datos[6][2] not in asig_4[datos[6][0]]:
                    asig_4[datos[6][0]][datos[6][2]] = list()
            else:
                if datos[6][2] not in asig_4[datos[6][0]]:
                    asig_4[datos[6][0]][datos[6][2]] = list()

#We populate the dictionary for 4, where there is a dictionary with the career name as the key, and inside another dictionary with the name of the mandatory courses for 4. Within each internal dictionary are the electives for that career.
for cur,cs in curso_opt.items():
    for cur2,dic in asig_4.items():
        for c2 in dic.keys():
            if cur.split("_")[0] == cur2:
                for c in cs:
                    asig_4[cur2][c2].append(c) 


#We form a new auxiliary dictionary where the electives are also included.
curso_comun2=curso_comun.copy()
curso_comun2.update(curso_opt)






# -------------------------We define the variables.------------------------

#It is a binary variable that indicates whether a subject is scheduled in a specific day, time slot, and classroom (it becomes 1)

x = LpVariable.dicts("clase_horario", ( dias, horarios,aula, asignaturas), cat='Binary')


# ----------------------------------Objective Function - Maximize preferences------------------------

#The function will be the variable x multiplied by the different preferences.
problem += lpSum((maña_tard[c][h] + preferencias_hora[p][h] + preferencias_dia[p]\
[int(dia)])* x[dia][h][a][c] for p, cs in profesores.items() for dia in dias\
 for c in cs for h in horarios for a in aula if c in asignaturas), "MaximizarPreferencias_hora"



#The second one is a binary variable that indicates if there is a class for a subject on a specific day and in a specific classroom. We will use it as an auxiliary variable.
clase_dia = LpVariable.dicts("EnseñanzaDia", (asignaturas, dias,aula), cat='Binary')

# --------------------Constraints---------------------------------
 

#The total number of classes for the subject will be the times it is taught per day multiplied by the number of days it is offered.
for c in asignaturas:
    problem += lpSum(x[dia][h][a][c] for h in horarios for dia in dias for a in aula) == dias_semana[c]*horas_dia[c]


#Each day, the subject will be taught either zero hours or the number of hours per day as specified in the Excel sheet.
for c in asignaturas:
    for dia in dias:
        problem += lpSum(clase_dia[c][dia][a] for a in aula) <= 1
        for a in aula:
            problem += lpSum(x[dia][h][a][c] for h in horarios) <= horas_dia[c] * lpSum(clase_dia[c][dia][a])
    problem += lpSum(clase_dia[c][dia][a] for a in aula for dia in dias) == dias_semana[c]


# Only one class can be taught per hour and classroom.
for h in horarios:
    for a in aula:
        for dia in dias:
            problem += lpSum(x[dia][h][a][c] for c in asignaturas ) <= 1


#There cannot be two subjects from the same course at the same time unless they are elective subjects from different mentions, or they are GA (General Assignments).
for cur, cs in curso_comun2.items():
    for dia in dias:
        for h in horarios:
            problem += lpSum(x[dia][h][a][c]  for c in cs for a in aula) <= 1


#For each course in each career, classes will be taught in the same classroom (excluding electives).
for dia in dias:
    a_1=0
    for cur,cd in curso.items():
        a_1 += 1
        for a in aula:
            if a_1!=a: #This ensures that if the class is not the one specified by the index, no class can be scheduled in that classroom.
                for c in cd:
                    problem += lpSum(x[dia][h][a][c]  for h in horarios) ==0  


#Elective courses that last more than one hour are scheduled consecutively.
for dia in dias:
    for cur,cd in curso_opt.items():
        for a in aula:
            for c  in cd:
                problem += lpSum(x[dia][h][a][c]  for h in horarios) == clase_dia[c][dia][a]*horas_dia[c]

            
# A professor cannot be in two places at the same time (teaching two classes simultaneously).
for pro, cs in profesores.items():
    for h in horarios:
        for dia in dias:
            problem += lpSum(x[dia][h][a][c] for c in cs if c in asignaturas for a in aula) <= 1


#The General Academics (GA) must occur at the same time as the main subjects.
for asi,cs in asignaturas_ga.items():
    for h in horarios:
        for dia in dias:
            problem += lpSum(x[dia][h][a][cs[1]]  for a in aula) <=  lpSum(x[dia][h][a][cs[0]] for a in aula)


#On Wednesdays, there are no classes from 11:30 AM to 3:00 PM.
for h in [4,5,6]:
    problem += lpSum(x[3][h][a][c]  for a in aula for c in asignaturas) == 0

#Elective courses cannot coincide with level 4 courses.
for car,dic in asig_4.items():
    for asig,cs in dic.items():
        for dia in dias:
            for h in horarios:
                for c in cs:
                    problem += lpSum(x[dia][h][a][asig] + x[dia][h][a][c] for a in aula) <=1



# ---------------------------------------Solving the problem.------------------------------
problem.solve()

# -----------------------We organize the results, keeping only the results where \( x = 1 \).----------------------------
print("Status:", LpStatus[problem.status])
datos_finales=[]
# To print the assignments (if necessary)
for v in problem.variables():
    if v.varValue==1:
        if v.name[0]=="c":
            
            datos_finales.append(v.name)

#-------------------------------------------------We transfer the data to Excel------------------------------------------------------------------------------
# Function to parse the data
def parsear_datos(linea):
    partes = re.match(r"clase_horario_(\d+)_(\d+)_(\d+)_(.*)", linea)
    if partes:
        return {
            "pagina": int(partes.group(1)),
            "dia": partes.group(1),
            "fila": int(partes.group(2)),
            "columna": int(partes.group(3)),
            "clase": partes.group(4).replace("_", " ")
        }
    return None

datos_parseados = [parsear_datos(d) for d in datos_finales]

# Defining colors and borders
fuente_negrita = Font(bold=True)
fondo_gris = PatternFill(start_color="00CCCCCC",
                         end_color="00CCCCCC",
                         fill_type="solid")
borde_delgado = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))

wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Mapping of page number to weekday
pagina_a_dia = {
    1: "Lunes",
    2: "Martes",
    3: "Miércoles",
    4: "Jueves",
    5: "Viernes"
}

# "Hours to insert on the left."
horas = ["8:40", "9:40", "10:40", "12:00", "13:00", "14:00", "15:00", "16:00", "17:00", "18:00"]

# "We organize the data by page."
datos_por_pagina = {}
for item in datos_parseados:
    if item:
        pagina = item["pagina"]
        if pagina not in datos_por_pagina:
            datos_por_pagina[pagina] = []
        datos_por_pagina[pagina].append(item)

# "We create the sheets and write the data."
for pagina, items in datos_por_pagina.items():
    dia = pagina_a_dia[pagina]  
    ws = wb.create_sheet(title=dia)
    
    # "We insert the class names in the first row."
    for i in range(len(aula)):  

        celda = ws.cell(row=1, column=i+2, value=f"Clase{i+1}" )
        celda.font = fuente_negrita
        celda.fill = fondo_gris
        celda.border = borde_delgado
    
    # "We insert hours in the first column."
    for i, hora in enumerate(horas, start=2):  
        celda= ws.cell(row=i, column=1, value=hora)
        celda.font = fuente_negrita
        celda.fill = fondo_gris
        celda.border = borde_delgado
    
    # "We write the class data, adjusting the rows and columns for the new headers."
    for item in items:
        fila = item["fila"] + 1  
        columna = item["columna"] + 1  
        clase = item["clase"]
        for j in profesores.keys():
            if clase in profesores[j]:
                profe = j

                break

        ws.cell(row=fila, column=columna, value=clase + " (" + profe+")")




# "We save the Excel file."
wb.save("Horario_PL_ga1.xlsx")
