#-------------------------------------------------This function converts the solution data into an Excel format-------------------------------------------------
#-------------------------------------------------Inputs: Excel file name and the solution in list form-----------------------------------------
#-------------------------------------------------Output: Creates an Excel file with the solution---------------------------------------------------------------


    
import pandas as pd
import numpy as np
import Lectura_de_datos as lect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side


def excel(solucion_semana,nombre):
    # Create a new Excel workbook


    aulas = 8 
    franjas = 10 
    horas_max = 4  
    horas_max_opt = 7 
    dias_semna = 5 
    wb = Workbook()
    hoja = wb.active
    
    horas_posibles=np.array(["8:40","9:40","10:40","12:00","13:00","14:00","15:00","16:00","17:00","18:00"])
    
    gris = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    border = Border(left=Side(border_style="thick"),
                    right=Side(border_style="thick"),
                    top=Side(border_style="thick"),
                    bottom=Side(border_style="thick"))

    # Iterate over the weekly solution and write data to the Excel sheet
    for dia, solucion_dia in enumerate(solucion_semana, start=1):  
        nombre_hoja = f"Dia {dia}"
        hoja = wb.create_sheet(title=nombre_hoja)

        # Write data to the sheet

        
        for clase in range(aulas):
            hoja.cell(row=1, column=clase+2, value="Clase " + str(clase+1))
            hoja.cell(row=1, column=clase + 2).font = Font(bold=True)
            hoja.cell(row=1, column=clase + 2).fill = gris
            hoja.cell(row=1, column=clase + 2).border = border

            for hora in range(franjas):
                hoja.cell(row=hora + 2, column=1, value=horas_posibles[hora])
                hoja.cell(row=hora + 2, column=1).font = Font(bold=True)
                hoja.cell(row=hora + 2, column=1).fill = gris
                hoja.cell(row=hora + 2 , column=1).border = border

                datos = solucion_dia[hora][clase]
                if datos is not None:
                    if isinstance(datos[6], np.ndarray):
                        elementos_texto = [str(elem) for elem in datos[6][2:3].tolist()]
                        profesor_texto = [str(elem2) for elem2 in datos[6][3:4].tolist()]
                        datos_texto = ', '.join(elementos_texto)+" ("+', '.join(profesor_texto)+")"
                    else:
                        
                        datos_texto = None

                    hoja.cell(row=hora+2, column=clase+2, value=datos_texto)

    wb.remove(wb['Sheet'])

    # Save the workbook to a file
    nombre_archivo = nombre + ".xlsx"
    wb.save(nombre_archivo)

    print(f"La solución de la semana se ha guardado en el archivo: {nombre_archivo}")